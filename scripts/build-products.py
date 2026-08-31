#!/usr/bin/env python3
"""
Build products.json from Excel catalog + existing template structure.
Reads PRODUCTOS (1).xlsx and generates the full product catalog.
"""
import json
import re
import openpyxl

# ─── Helpers ──────────────────────────────────────────────────────────────

def slugify(text: str) -> str:
    """Convert text to URL-safe slug."""
    t = text.lower().strip()
    t = re.sub(r'[áàäâ]', 'a', t)
    t = re.sub(r'[éèëê]', 'e', t)
    t = re.sub(r'[íìïî]', 'i', t)
    t = re.sub(r'[óòöô]', 'o', t)
    t = re.sub(r'[úùüû]', 'u', t)
    t = re.sub(r'[ñ]', 'n', t)
    t = re.sub(r'[^a-z0-9]+', '-', t)
    t = t.strip('-')
    return t

def extract_measure(name: str) -> str:
    """Extract dimensions from product name like 'Cajas de Cartón (40x60x40)'."""
    m = re.search(r'\(([^)]+?)\)', name)
    if m:
        return m.group(1).strip().replace('X', 'x').replace('×', 'x')
    # Try patterns like "24x35x35"
    m = re.search(r'(\d+[\.\d]*\s*[xX×]\s*\d+[\.\d]*\s*[xX×]\s*\d+[\.\d]*)', name)
    if m:
        return m.group(1).strip().replace('X', 'x').replace('×', 'x')
    return ""

def make_packs(base_price: float) -> list:
    """Generate 3 pack tiers with progressive discounts."""
    return [
        {"cantidad": 25, "precio": round(base_price * 25 * 0.97, 2), "descuento": 3},
        {"cantidad": 50, "precio": round(base_price * 50 * 0.93, 2), "descuento": 7},
        {"cantidad": 100, "precio": round(base_price * 100 * 0.88, 2), "descuento": 12},
    ]

def make_product(code: str, name: str, categoria: str, subcategoria: str,
                  medida: str, short_desc: str = "", full_desc: str = "",
                  excel_price: float = 0, material: str = "",
                  colores: list = None, etiquetas: list = None,
                  destacado: bool = False) -> dict:
    """Build a product object matching the existing schema."""
    if not colores:
        colores = ["Natural Kraft"]
    if not etiquetas:
        etiquetas = []

    clean_name = name.strip()
    base_price = excel_price if excel_price > 0 else estimate_price(subcategoria, medida, clean_name)

    # Build slug from code
    slug_base = slugify(clean_name)
    slug = f"{slug_base}" if slug_base else slugify(code)

    # Generate image paths
    img_base = slug.replace('-', '-')[:50]
    imagenes = [
        f"/images/products/{img_base}-1.jpg",
        f"/images/products/{img_base}-2.jpg",
        f"/images/products/{img_base}-3.jpg",
    ]

    # Build description
    if full_desc:
        # Clean markdown-like headers
        desc = re.sub(r'\[H[1-3]\]\s*', '', full_desc).strip()
        if len(desc) < 50:
            desc = build_description(clean_name, subcategoria, medida, material)
    elif short_desc:
        desc = short_desc.strip()
        if len(desc) < 50:
            desc = build_description(clean_name, subcategoria, medida, material)
    else:
        desc = build_description(clean_name, subcategoria, medida, material)

    return {
        "id": slugify(code) if code else slug,
        "nombre": clean_name,
        "descripcion": desc,
        "precio": round(base_price, 2),
        "precioUnidad": f"S/ {base_price:.2f}",
        "stock": 500,
        "categoria": categoria,
        "material": material or guess_material(subcategoria, categoria),
        "imagenes": imagenes,
        "colores": colores,
        "medidas": [medida] if medida else [],
        "destacado": destacado,
        "etiquetas": etiquetas,
        "packs": make_packs(base_price),
        "slug": slug,
    }

def estimate_price(subcategoria: str, medida: str, name: str = "") -> float:
    """Estimate a reasonable unit price based on product type and name details."""
    n = name.lower()
    # Normalize accents for matching
    n = re.sub(r'[áàäâ]', 'a', n)
    n = re.sub(r'[éèëê]', 'e', n)
    n = re.sub(r'[íìïî]', 'i', n)
    n = re.sub(r'[óòöô]', 'o', n)
    n = re.sub(r'[úùüû]', 'u', n)
    n = re.sub(r'[ñ]', 'n', n)
    # Extract a volume hint from measure
    nums = re.findall(r'[\d.]+', medida)
    vol = 1
    if len(nums) >= 3:
        vol = float(nums[0]) * float(nums[1]) * float(nums[2]) / 1000
    elif len(nums) >= 2:
        vol = float(nums[0]) * float(nums[1]) / 100

    # Stretch film — use weight from name (e.g. "0.50 KG", "2 KG")
    if 'stretch' in n:
        kg_match = re.search(r'(\d+[\.\d]*)\s*KG', n, re.IGNORECASE)
        width_match = re.search(r'(\d+)\s*[″"]', n)
        kg = float(kg_match.group(1)) if kg_match else 1.0
        width = int(width_match.group(1)) if width_match else 20
        # Base: ~S/6 per kg for 20" roll, adjust for width
        price_per_kg = 6.0 * (width / 20)
        return round(kg * price_per_kg, 2)

    if 'archivera' in subcategoria.lower() or 'archivera' in n:
        return round(max(3.5, vol * 0.4 + 1.5), 2)
    if 'doble corrugado' in subcategoria.lower() or 'doble' in subcategoria.lower() or 'doble' in n:
        return round(max(5.0, vol * 0.5 + 2.0), 2)
    if 'pizza' in subcategoria.lower() or 'pizza' in n:
        nums2 = re.findall(r'[\d.]+', name)
        # Use the first dimension as the side
        side = 20
        for num in nums2:
            v = float(num)
            if v >= 10 and v <= 50:
                side = v
                break
        return round(max(1.50, side * 0.008 + 1.0), 2)
    if 'postal' in subcategoria.lower() or 'postal' in n or 'envio' in n:
        return round(max(1.50, vol * 0.25 + 0.8), 2)
    if 'navidena' in subcategoria.lower() or 'navidena' in n:
        return round(max(8.0, vol * 0.3 + 5.0), 2)
    if 'cinta' in subcategoria.lower() or 'cinta' in n:
        # Cinta price based on yards and microns
        yds_match = re.search(r'(\d+)\s*YDS', n, re.IGNORECASE)
        mc_match = re.search(r'(\d+)\s*MC', n, re.IGNORECASE)
        yds = int(yds_match.group(1)) if yds_match else 180
        mc = int(mc_match.group(1)) if mc_match else 40
        price_per_yd = (mc / 40) * 0.03
        return round(yds * price_per_yd + 1.5, 2)
    if 'burbuja' in subcategoria.lower() or 'burbupack' in n or 'burbuja' in n:
        cm_match = re.search(r'(\d+)\s*CM', n, re.IGNORECASE)
        cm = int(cm_match.group(1)) if cm_match else 100
        return round(cm * 0.04 + 1.0, 2)
    if 'burbupack' in n:
        cm_match = re.search(r'(\d+)', n)
        cm = int(cm_match.group(1)) if cm_match else 100
        return round(cm * 0.04 + 1.0, 2)
    if 'tecnopor' in n:
        return round(20.0, 2)
    if 'plancha' in n and 'doble' in n:
        return round(16.0, 2)
    if 'plancha' in n and 'micro' in n:
        return round(7.0, 2)
    if 'plancha' in n:
        return round(11.0, 2)
    if 'papel' in n and 'embalar' in n:
        return round(8.50, 2)
    if 'viruta' in n and 'natural' in n:
        return round(30.0, 2)
    if 'viruta' in n:
        return round(25.0, 2)
    if 'carton corrugado' in n and 'cm' in n:
        cm_match = re.search(r'(\d+)\s*CM', n, re.IGNORECASE)
        cm = int(cm_match.group(1)) if cm_match else 100
        return round(cm * 0.06 + 2.0, 2)
    # Default
    return round(max(3.0, vol * 0.3 + 1.5), 2)

def guess_material(subcategoria: str, categoria: str) -> str:
    """Guess material from subcategory."""
    sc = subcategoria.lower()
    if 'doble corrugado' in sc or 'doble' in sc:
        return "Cartón Doble Corrugado"
    if 'archivera' in sc:
        return "Cartón Corrugado Archivero"
    if 'pizza' in sc:
        return "Cartón Corrugado Alimentario"
    if 'postal' in sc or 'envío' in sc:
        return "Cartón Corrugado E-commerce"
    if 'navideña' in sc:
        return "Cartón Corrugado Reforzado"
    if 'stretch' in sc or 'film' in sc:
        return "Film Estirable PE"
    if 'cinta' in sc:
        return "Polipropileno"
    if 'burbuja' in sc or 'burbupack' in sc:
        return "Polietileno Burbuja"
    if 'plancha' in sc and 'tecnopor' in sc:
        return "Poliestireno Expandido (Tecnopor)"
    if 'plancha' in sc and ('doble' in sc or 'corrugado' in sc):
        return "Cartón Doble Corrugado"
    if 'plancha' in sc:
        return "Cartón Corrugado"
    if 'papel' in sc:
        return "Papel Kraft"
    if 'viruta' in sc:
        return "Viruta de Papel Reciclado"
    if 'cartón' in sc:
        return "Cartón Corrugado"
    return "Cartón Corrugado"

def build_description(name: str, subcat: str, medida: str, material: str) -> str:
    """Build a professional product description."""
    parts = []
    if 'archivera' in subcat.lower():
        parts.append(f"{name}. Cajas archiveras de alta calidad ideales para organización de oficinas, archivos y documentos importantes.")
    elif 'pizza' in subcat.lower():
        parts.append(f"{name}. Cajas diseñadas específicamente para el servicio de delivery de pizza y alimentos calientes.")
    elif 'postal' in subcat.lower() or 'envío' in subcat.lower():
        parts.append(f"{name}. Cajas de envío resistentes diseñadas para ecommerce y logística de paquetería.")
    elif 'doble' in subcat.lower():
        parts.append(f"{name}. Caja de doble corrugado para máxima resistencia en envíos pesados e industriales.")
    elif 'navideña' in subcat.lower():
        parts.append(f"{name}. Caja reforzada especial para canastas navideñas corporativas.")
    else:
        parts.append(f"{name}. Producto de embalaje profesional para uso comercial e industrial.")

    if medida:
        parts.append(f"Medidas exteriores: {medida} cm.")
    if material:
        parts.append(f"Material: {material}.")
    parts.append("Fabricado con materiales de primera calidad. Venta por mayor con descuentos progresivos. Ideal para empresas, tiendas online y logística.")
    return " ".join(parts)


# ─── Read Excel ───────────────────────────────────────────────────────────

wb = openpyxl.load_workbook('/home/z/my-project/upload/PRODUCTOS (1).xlsx', data_only=True)
ws = wb['Hoja1']

# ─── Parse hierarchical structure ─────────────────────────────────────────

CATEGORY_MAP = {
    "cajas archiveras": "Cajas",
    "cajas de cartón": "Cajas",
    "cajas e-commerce": "Cajas",
    "cajas navideñas": "Cajas",
    "stretch film": "Films",
    "cintas de embalaje": "Cintas",
    "relleno y protección": "Protección",
}

# Hoja2: Carton types reference (for material descriptions)
hoja2 = {}
ws2 = wb['Hoja2']
for r in range(2, ws2.max_row + 1):
    code = ws2.cell(r, 2).value
    desc = ws2.cell(r, 3).value
    if code and desc:
        hoja2[str(code).strip()] = str(desc).strip()

products = []
current_categoria_header = ""
current_subcategoria = ""
current_categoria = ""
product_count = 0

for r in range(1, ws.max_row + 1):
    col_b = ws.cell(r, 2).value  # Code
    col_c = ws.cell(r, 3).value  # Category header
    col_d = ws.cell(r, 4).value  # Subcategory / Product name
    col_f = ws.cell(r, 6).value  # Short description
    col_g = ws.cell(r, 7).value  # Full description
    col_i = ws.cell(r, 9).value  # Price (sometimes)
    col_h = ws.cell(r, 8).value  # Extra info

    # Detect category header (col C)
    if col_c and col_c != 'CATEGORIA':
        current_categoria_header = str(col_c).strip()
        # Map to our 5 categories
        current_categoria = "Cajas"  # default
        for key, val in CATEGORY_MAP.items():
            if key in current_categoria_header.lower():
                current_categoria = val
                break
        current_subcategoria = ""
        continue

    # Detect subcategory (col D, no code in col B)
    if col_d and not col_b:
        current_subcategoria = str(col_d).strip()
        continue

    # Detect product (has code in col B and name in col D)
    if col_b and col_d:
        code = str(col_b).strip()
        name = str(col_d).strip()
        medida = extract_measure(name)

        # Price from Excel if available
        excel_price = 0
        if col_i and isinstance(col_i, (int, float)):
            excel_price = float(col_i)

        short_desc = str(col_f).strip() if col_f else ""
        full_desc = str(col_g).strip() if col_g else ""

        # Determine colors based on product type
        colores = ["Natural Kraft"]
        if 'blanco' in name.lower() or 'blanco' in current_subcategoria.lower():
            colores = ["Blanco", "Natural Kraft"]
        elif 'pizza' in current_subcategoria.lower():
            colores = ["Natural Kraft", "Blanco"]
        elif 'stretch' in current_categoria_header.lower() and 'negro' in name.lower():
            colores = ["Negro"]
        elif 'cinta' in current_categoria_header.lower():
            colores = ["Transparente"]
            if 'fragil' in name.lower():
                colores = ["Transparente con impresión"]

        # Determine tags
        etiquetas = []
        sc_lower = current_subcategoria.lower()
        name_lower = name.lower()
        if 'archivera' in sc_lower:
            etiquetas = ["archivo", "oficina", "organización"]
        elif 'doble' in sc_lower:
            etiquetas = ["doble pared", "reforzado", "industrial"]
        elif 'pizza' in sc_lower:
            etiquetas = ["alimentario", "delivery", "caliente"]
        elif 'postal' in sc_lower or 'envío' in sc_lower:
            etiquetas = ["ecommerce", "envío", "correría"]
        elif 'navideña' in sc_lower:
            etiquetas = ["navideño", "corporativo", "regalo"]
        elif 'stretch' in current_categoria_header.lower():
            if 'negro' in name_lower:
                etiquetas = ["opaco", "protección", "paletizado"]
            else:
                etiquetas = ["transparente", "paletizado", "resistente"]
        elif 'cinta' in current_categoria_header.lower():
            if 'fragil' in name_lower:
                etiquetas = ["fragil", "seguridad", "advertencia"]
            else:
                etiquetas = ["cierre", "empaque", "resistente"]
        elif 'burbuja' in sc_lower or 'burbupack' in name_lower:
            etiquetas = ["protección", "relleno", "fragil"]
        elif 'tecnopor' in name_lower:
            etiquetas = ["aislante", "protección", "relleno"]
        elif 'papel' in name_lower and 'embalar' in name_lower:
            etiquetas = ["ecológico", "relleno", "protección"]
        elif 'viruta' in name_lower:
            etiquetas = ["ecológico", "relleno", "presentación"]
        elif 'plancha' in name_lower:
            etiquetas = ["personalizable", "corte", "industrial"]
        else:
            etiquetas = ["professional", "resistente"]

        # Featured: first product in each subcategory
        destacado = (product_count % 8 == 0)

        # Build material string
        mat = guess_material(current_subcategoria, current_categoria)

        # For stretch film, extract width/weight from name
        if 'stretch' in current_categoria_header.lower():
            mat = "Film Estirable Polietileno"
            if 'negro' in name_lower:
                mat = "Film Estirable Polietileno Negro"

        # For cintas
        if 'cinta' in current_categoria_header.lower():
            mat = "Cinta Adhesiva de Polipropileno"
            if 'fragil' in name_lower:
                mat = "Cinta Adhesiva de Polipropileno con Impresión Frágil"

        p = make_product(
            code=code,
            name=name,
            categoria=current_categoria,
            subcategoria=current_subcategoria,
            medida=medida,
            short_desc=short_desc,
            full_desc=full_desc,
            excel_price=excel_price,
            material=mat,
            colores=colores,
            etiquetas=etiquetas,
            destacado=destacado,
        )
        products.append(p)
        product_count += 1

# ─── Output ───────────────────────────────────────────────────────────────

print(f"Generated {len(products)} products")
print(f"Categories: {set(p['categoria'] for p in products)}")
print(f"Featured: {sum(1 for p in products if p['destacado'])}")

# Verify no duplicate slugs
slugs = [p['slug'] for p in products]
dupes = [s for s in slugs if slugs.count(s) > 1]
if dupes:
    print(f"WARNING: Duplicate slugs: {set(dupes)}")
    # Fix duplicates
    seen = {}
    for p in products:
        if p['slug'] in seen:
            base = p['slug']
            i = 2
            while f"{base}-{i}" in seen:
                i += 1
            p['slug'] = f"{base}-{i}"
            p['id'] = p['slug']
        seen[p['slug']] = True

# Verify no duplicate IDs
ids = [p['id'] for p in products]
dupes_id = [i for i in ids if ids.count(i) > 1]
if dupes_id:
    print(f"WARNING: Duplicate IDs: {set(dupes_id)}")

# Write
with open('/home/z/my-project/src/data/products.json', 'w', encoding='utf-8') as f:
    json.dump(products, f, ensure_ascii=False, indent=2)

print("✅ Written to /home/z/my-project/src/data/products.json")

# Print summary
for p in products:
    print(f"  {p['id'][:45]:45s} | {p['categoria']:12s} | {p['nombre'][:50]} | S/{p['precio']:.2f}")